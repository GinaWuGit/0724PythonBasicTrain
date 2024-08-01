from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
#1. 讀取檔案
wb = load_workbook('22CH0X.xlsx')
ws = wb.active                     #讀取預設工作表
#ws = wb['表2 Inverter Error Code']  #讀取指定工作表
print(ws)
print(ws['F3'].value)

#2. 修正cell內容
# ws['F3'].value = 'Main System Power(Low Word)'
# wb.save('22CH0X.xlsx')

#3. 印出所有工作表
# print(wb.sheetnames)

#4. 新增sheet
# wb.create_sheet('Modbus')
# wb.save('22CH0X.xlsx')

#5. 讀取範圍資料
for row in range (3, 6):
    for col in range (6, 11):
        char = get_column_letter(col)
        print(ws[char + str(row)].value)